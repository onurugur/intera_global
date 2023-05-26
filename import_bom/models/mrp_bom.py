# -*- coding: utf-8 -*-

from odoo import api, fields, models, _
from odoo.tools.float_utils import float_is_zero, float_compare
from odoo.exceptions import ValidationError, UserError
from odoo.tools import DEFAULT_SERVER_DATETIME_FORMAT
import datetime

from io import BytesIO
import odoo.tools as tools
import openpyxl
import base64
import re


class WizardImportBom(models.TransientModel):
    _name = 'wizard.import.bom'
    _description = 'Import Bom Lines'

    name = fields.Binary('File')
    msg = fields.Text('Messages', readonly=True)
    delete_old_values=fields.Boolean('Delete old Values',default=False)

    
    def import_bom(self):
        fdata = self.name and base64.decodestring(self.name) or False
        msg = ''
        bom_line_obj = self.env['mrp.bom.line']
        bom_obj=self.env['mrp.bom']
        product_obj = self.env['product.product']
        operation_obj=self.env['mrp.routing.workcenter']
        input = BytesIO(fdata)
        bom=bom_obj.browse(self._context.get('active_id'))
        if self.delete_old_values:
            bom.bom_line_ids.unlink()
        input.seek(0)
        wb = openpyxl.load_workbook(input)
        sheet = wb._sheets[0]
        operation_id=False
        #upgrade fix
#         if bom_obj.browse(self._context.get('active_id')).state == 'release':
#             raise UserError(_("You can not import to released Bom"))
#         else:
        for rowNum in range(2, sheet.max_row+1):
                if sheet.cell(row=rowNum, column=1).value:
                    product_name = str(sheet.cell(row=rowNum, column=1).value)
                else:
                    continue
                product = product_obj.search([('default_code', '=', product_name)],limit=1)
                if not product:
                    raise UserError(_('%s is not in products' %(product_name)))
                if sheet.max_column >= 2:
                    product_qty = float(sheet.cell(row=rowNum, column=2).value)
                else:
                    product_qty = 1
                if sheet.max_column >=3 :
                    operations=bom.operation_ids.mapped('name')
                    operation_id=operation_obj.search(
                        [('name','=',str(sheet.cell(row=rowNum, column=3).value)),('bom_id','=',bom.id)],limit=1
                        ).filtered(lambda l:str(sheet.cell(row=rowNum, column=1).value  in operations))
                bom_line_obj.create({
                    'bom_id': self._context.get('active_id'),
                    'product_id': product.id,
                    'product_qty': product_qty,
                    'operation_id':operation_id and operation_id.id,
                })

